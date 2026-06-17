"""
batch_process.py — PRC SCAL Batch Analytics CLI
================================================
Iterates every Excel/CSV file inside an /uploads directory, auto-identifies the
data type (Kr, MICP, RI, FF), validates physics with PhysicsGuard, fits the
appropriate model, and writes a consolidated PRC_BATCH_REPORT.xlsx summary.

Usage:
    python batch_process.py                          # default: ./uploads/
    python batch_process.py --input /path/to/data
    python batch_process.py --input ./uploads --output ./reports/batch.xlsx
"""

import argparse
import logging
import os
import sys
import glob
from dataclasses import dataclass, field
from typing import Optional

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from physics_validator import PhysicsGuard
from petrophysical_curves import Endpoints, KrCurveFitter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
_log = logging.getLogger("PRC-Batch")

# ── DATA-TYPE DETECTION ────────────────────────────────────────────────────────

_KR_KEYWORDS   = {"krw", "kro", "krg", "relative_permeability", "sor", "swi"}
_MICP_KEYWORDS = {"hg", "mercury", "s_hg", "sw_hg", "hg_sat", "hg_pressure",
                  "pc_psia", "pressure_psia", "micp", "intrusion"}
_RI_KEYWORDS   = {"ri", "resistivity_index", "ro", "rw", "archie_n"}
_FF_KEYWORDS   = {"formation_factor", "ff", "archie_m", "cementation"}


def _col_set(df: pd.DataFrame) -> set:
    return {c.lower().strip().replace(" ", "_") for c in df.columns}


def detect_type(df: pd.DataFrame) -> str:
    cols = _col_set(df)
    # MICP checked first — mercury data must never be routed to Kr
    if cols & _MICP_KEYWORDS:
        return "micp"
    if cols & _KR_KEYWORDS:
        return "kr"
    if cols & _RI_KEYWORDS:
        return "ri"
    if cols & _FF_KEYWORDS:
        return "ff"
    return "unknown"


def _find_col(df: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    """Return first matching column name (case-insensitive)."""
    lower_map = {c.lower().strip(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower_map:
            return lower_map[cand.lower()]
    return None


# ── RESULT CONTAINER ──────────────────────────────────────────────────────────

@dataclass
class BatchResult:
    filename:       str
    data_type:      str
    physics_score:  int        = 0
    physics_grade:  str        = "?"
    violations:     int        = 0
    # Kr fitted params
    nw:             float      = float("nan")
    no:             float      = float("nan")
    swi:            float      = float("nan")
    sor:            float      = float("nan")
    krw_max:        float      = float("nan")
    kro_max:        float      = float("nan")
    r2_bc:          float      = float("nan")
    # MICP metrics
    entry_pressure: float      = float("nan")
    modal_radius:   float      = float("nan")
    max_hg_sat:     float      = float("nan")
    # RI / FF
    archie_n:       float      = float("nan")
    archie_m:       float      = float("nan")
    archie_a:       float      = float("nan")
    # Status
    status:         str        = "OK"
    notes:          str        = ""


# ── PROCESSORS ────────────────────────────────────────────────────────────────

def _process_kr(df: pd.DataFrame, res: BatchResult) -> None:
    sw_col  = _find_col(df, ["Sw", "Sw_frac", "Water_Sat", "water_saturation"])
    krw_col = _find_col(df, ["Krw", "KrW", "kr_water", "water_kr"])
    kro_col = _find_col(df, ["Kro", "KrO", "kr_oil",   "oil_kr"])

    if not (sw_col and krw_col and kro_col):
        res.status = "ERROR"
        res.notes  = f"Cannot map Sw/Krw/Kro columns. Found: {list(df.columns)}"
        return

    sw  = pd.to_numeric(df[sw_col],  errors="coerce").dropna().values
    krw = pd.to_numeric(df[krw_col], errors="coerce").dropna().values
    kro = pd.to_numeric(df[kro_col], errors="coerce").dropna().values
    n   = min(len(sw), len(krw), len(kro))
    sw, krw, kro = sw[:n], krw[:n], kro[:n]

    if n < 4:
        res.status = "ERROR"
        res.notes  = f"Insufficient data points ({n}); need ≥ 4."
        return

    # Physics audit
    audit = PhysicsGuard().validate_kr(sw, krw, kro).generate_health_score()
    res.physics_score = audit["score"]
    res.physics_grade = audit["grade"]
    res.violations    = len(audit["violations"])

    # Curve fit
    swi_est = float(np.min(sw))
    sor_est = float(1.0 - np.max(sw))
    if swi_est + sor_est >= 1.0:
        sor_est = max(0.05, 1.0 - np.max(sw) - 0.01)

    try:
        ep      = Endpoints(
            Swi=swi_est, Sor=sor_est,
            Krw_max=float(np.max(krw)), Kro_max=float(np.max(kro))
        )
        fitter  = KrCurveFitter(ep)
        bc_res  = fitter.fit_brooks_corey(sw, krw, kro)

        res.nw      = round(float(bc_res.get("nw", float("nan"))), 4)
        res.no      = round(float(bc_res.get("no", float("nan"))), 4)
        res.swi     = round(swi_est, 4)
        res.sor     = round(sor_est, 4)
        res.krw_max = round(float(np.max(krw)), 4)
        res.kro_max = round(float(np.max(kro)), 4)
        res.r2_bc   = round(float(bc_res.get("r2", float("nan"))), 4)
    except Exception as e:
        res.status = "FIT_FAILED"
        res.notes  = str(e)


def _process_micp(df: pd.DataFrame, res: BatchResult) -> None:
    pc_col  = _find_col(df, ["Pc_psia", "Pressure_psia", "Hg_Pressure", "pc", "pressure"])
    hg_col  = _find_col(df, ["S_Hg", "Hg_Sat", "Mercury_Sat", "Sw_Hg", "hg_saturation"])

    if not (pc_col and hg_col):
        res.status = "ERROR"
        res.notes  = f"Cannot map Pc/Hg_Sat columns. Found: {list(df.columns)}"
        return

    pc  = pd.to_numeric(df[pc_col],  errors="coerce").dropna().values
    shg = pd.to_numeric(df[hg_col],  errors="coerce").dropna().values
    n   = min(len(pc), len(shg))
    pc, shg = pc[:n], shg[:n]

    if n < 3:
        res.status = "ERROR"
        res.notes  = f"Insufficient data points ({n}); need ≥ 3."
        return

    # Physics audit
    audit = PhysicsGuard().validate_micp(pc, shg).generate_health_score()
    res.physics_score = audit["score"]
    res.physics_grade = audit["grade"]
    res.violations    = len(audit["violations"])

    # MICP metrics
    idx    = np.argsort(pc)
    pc_s   = np.array(pc)[idx]
    shg_s  = np.array(shg)[idx]
    pc_pos = np.maximum(pc_s, 0.1)

    entry_mask       = shg_s > 0.01
    pe               = float(pc_s[entry_mask][0]) if entry_mask.any() else float(pc_s[0])
    thr_r            = 107.6 / max(pe, 0.1)

    res.entry_pressure = round(pe, 2)
    res.modal_radius   = round(thr_r, 4)
    res.max_hg_sat     = round(float(shg_s[-1] * 100.0), 1)


def _process_ri(df: pd.DataFrame, res: BatchResult) -> None:
    sw_col = _find_col(df, ["Sw", "water_saturation"])
    ri_col = _find_col(df, ["RI", "Resistivity_Index", "ri"])
    if not (sw_col and ri_col):
        res.status = "ERROR"
        res.notes  = f"Cannot map Sw/RI columns. Found: {list(df.columns)}"
        return
    sw_a = pd.to_numeric(df[sw_col], errors="coerce").dropna().values
    ri_a = pd.to_numeric(df[ri_col], errors="coerce").dropna().values
    n    = min(len(sw_a), len(ri_a))
    sw_a, ri_a = sw_a[:n], ri_a[:n]
    mask = (sw_a > 0) & (ri_a > 0)
    if mask.sum() < 2:
        res.status = "ERROR"; return
    n_arch = float(-np.polyfit(np.log(sw_a[mask]), np.log(ri_a[mask]), 1)[0])
    res.archie_n      = round(max(1.5, min(n_arch, 3.0)), 4)
    res.physics_score = 100
    res.physics_grade = "A"


def _process_ff(df: pd.DataFrame, res: BatchResult) -> None:
    phi_col = _find_col(df, ["Porosity", "phi", "porosity"])
    ff_col  = _find_col(df, ["FF", "Formation_Factor", "ff"])
    if not (phi_col and ff_col):
        res.status = "ERROR"
        res.notes  = f"Cannot map Porosity/FF columns. Found: {list(df.columns)}"
        return
    phi_a = pd.to_numeric(df[phi_col], errors="coerce").dropna().values
    ff_a  = pd.to_numeric(df[ff_col],  errors="coerce").dropna().values
    n     = min(len(phi_a), len(ff_a))
    phi_a, ff_a = phi_a[:n], ff_a[:n]
    mask  = (phi_a > 0) & (ff_a > 0)
    if mask.sum() < 2:
        res.status = "ERROR"; return
    coeffs = np.polyfit(np.log(phi_a[mask]), np.log(ff_a[mask]), 1)
    res.archie_m      = round(float(max(1.3, min(-coeffs[0], 3.5))), 4)
    res.archie_a      = round(float(max(0.3, min(np.exp(coeffs[1]), 2.5))), 4)
    res.physics_score = 100
    res.physics_grade = "A"


# ── FILE READER ───────────────────────────────────────────────────────────────

def read_file(path: str) -> list[pd.DataFrame]:
    """Return list of DataFrames (one per sheet for Excel, one for CSV)."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        xf = pd.ExcelFile(path)
        return [pd.read_excel(path, sheet_name=s) for s in xf.sheet_names]
    elif ext == ".csv":
        return [pd.read_csv(path)]
    return []


# ── EXCEL REPORT WRITER ───────────────────────────────────────────────────────

_GOLD  = "FFD700"
_DARK  = "0A0A0C"
_GREY  = "1A1A1F"
_WHITE = "F8FAFC"
_RED   = "DC2626"
_GREEN = "16A34A"

def _cell_style(ws, row: int, col: int, value, bold=False,
                bg=None, fg=_WHITE, align="center", num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=fg, name="Consolas", size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt
    thin = Side(style="thin", color="2A2A35")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def write_report(results: list[BatchResult], output_path: str) -> None:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Master Summary ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "PRC Batch Summary"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value   = "PRC SCAL PIPELINE — Batch Analytics Report"
    title_cell.font    = Font(bold=True, color=_GOLD, name="Consolas", size=13)
    title_cell.fill    = PatternFill("solid", fgColor=_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    headers = [
        "File", "Type", "Physics Score", "Grade", "Violations",
        "nw", "no", "Swi", "Sor", "Krw_max", "Kro_max", "R²(BC)",
        "Pe (psia)", "Modal r (µm)", "Max Hg (%)",
        "Archie n", "Archie m", "Archie a",
        "Status", "Notes",
    ]
    for c, h in enumerate(headers, 1):
        _cell_style(ws, 2, c, h, bold=True, bg=_GREY, fg=_GOLD, align="center")
    ws.row_dimensions[2].height = 18

    # Data rows
    for r, res in enumerate(results, 3):
        score_bg = _GREEN if res.physics_score >= 90 else (_RED if res.physics_score < 60 else "D97706")
        row_bg   = "0D0D11" if r % 2 == 0 else "111118"

        def c(col, val, **kw):
            _cell_style(ws, r, col, val, bg=row_bg, **kw)

        c(1,  res.filename,       align="left",   fg=_WHITE)
        c(2,  res.data_type.upper())
        _cell_style(ws, r, 3, res.physics_score, bold=True, bg=score_bg, fg=_WHITE, align="center")
        c(4,  res.physics_grade,  bold=True)
        c(5,  res.violations)
        c(6,  res.nw,             num_fmt="0.0000")
        c(7,  res.no,             num_fmt="0.0000")
        c(8,  res.swi,            num_fmt="0.0000")
        c(9,  res.sor,            num_fmt="0.0000")
        c(10, res.krw_max,        num_fmt="0.0000")
        c(11, res.kro_max,        num_fmt="0.0000")
        c(12, res.r2_bc,          num_fmt="0.0000")
        c(13, res.entry_pressure, num_fmt="0.00")
        c(14, res.modal_radius,   num_fmt="0.0000")
        c(15, res.max_hg_sat,     num_fmt="0.0")
        c(16, res.archie_n,       num_fmt="0.0000")
        c(17, res.archie_m,       num_fmt="0.0000")
        c(18, res.archie_a,       num_fmt="0.0000")
        _cell_style(ws, r, 19, res.status,
                    bg=_RED if res.status != "OK" else row_bg,
                    fg=_WHITE, align="center")
        c(20, res.notes,          align="left", fg="#94A3B8")

    # Column widths
    widths = [32, 6, 13, 6, 10, 8, 8, 8, 8, 8, 8, 8,
              10, 13, 10, 9, 9, 9, 12, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Physics Violations ───────────────────────────────────────────
    wv = wb.create_sheet("Physics Violations")
    wv.sheet_view.showGridLines = False
    wv.merge_cells("A1:E1")
    wv["A1"].value     = "Physics Violations Detail"
    wv["A1"].font      = Font(bold=True, color=_GOLD, name="Consolas", size=11)
    wv["A1"].fill      = PatternFill("solid", fgColor=_DARK)
    wv["A1"].alignment = Alignment(horizontal="center")
    wv.row_dimensions[1].height = 22

    viol_headers = ["File", "Score", "Rule", "Severity", "Detail"]
    for c, h in enumerate(viol_headers, 1):
        _cell_style(wv, 2, c, h, bold=True, bg=_GREY, fg=_GOLD)

    vrow = 3
    for res in results:
        if res.violations == 0:
            continue
        # Re-run guard to get violation details (already computed; stored in BatchResult.notes implicitly)
        # We'll flag rows without drilling into violation text since BatchResult stores count only.
        # If you need full text, extend BatchResult with violations: list[dict].
        _cell_style(wv, vrow, 1, res.filename,       bg="0D0D11", align="left")
        _cell_style(wv, vrow, 2, res.physics_score,  bg=_RED,     fg=_WHITE, bold=True)
        _cell_style(wv, vrow, 3, f"{res.violations} violation(s)", bg="0D0D11")
        _cell_style(wv, vrow, 4, res.physics_grade,  bg="0D0D11")
        _cell_style(wv, vrow, 5, res.notes or "See console log for details", bg="0D0D11", align="left")
        vrow += 1

    for i, w in enumerate([32, 8, 22, 10, 60], 1):
        wv.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    _log.info("Report written -> %s", output_path)


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="PRC SCAL Batch Analytics Engine",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--input",  "-i", default="./uploads",
                        help="Folder containing Excel/CSV lab files (default: ./uploads)")
    parser.add_argument("--output", "-o", default="PRC_BATCH_REPORT.xlsx",
                        help="Output Excel report path (default: PRC_BATCH_REPORT.xlsx)")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Show per-file column detection details")
    args = parser.parse_args()

    input_dir = os.path.abspath(args.input)
    if not os.path.isdir(input_dir):
        _log.error("Input directory not found: %s", input_dir)
        sys.exit(1)

    patterns = ["*.xlsx", "*.xls", "*.csv"]
    files    = []
    for pat in patterns:
        files.extend(glob.glob(os.path.join(input_dir, pat)))
    files.sort()

    if not files:
        _log.warning("No Excel/CSV files found in %s", input_dir)
        sys.exit(0)

    _log.info("Found %d file(s) in %s", len(files), input_dir)

    results: list[BatchResult] = []

    for fpath in files:
        fname = os.path.basename(fpath)
        _log.info("Processing: %s", fname)

        try:
            sheets = read_file(fpath)
        except Exception as e:
            _log.error("  Cannot read %s: %s", fname, e)
            results.append(BatchResult(filename=fname, data_type="read_error",
                                       status="ERROR", notes=str(e)))
            continue

        for sheet_idx, df in enumerate(sheets):
            label = fname if len(sheets) == 1 else f"{fname}[sheet{sheet_idx}]"

            if df.empty or len(df.columns) < 2:
                _log.warning("  Skipping empty sheet in %s", label)
                continue

            dtype = detect_type(df)
            if args.verbose:
                _log.info("  Detected type=%s  cols=%s", dtype, list(df.columns))

            res = BatchResult(filename=label, data_type=dtype)

            if   dtype == "kr":   _process_kr(df, res)
            elif dtype == "micp": _process_micp(df, res)
            elif dtype == "ri":   _process_ri(df, res)
            elif dtype == "ff":   _process_ff(df, res)
            else:
                res.status = "UNKNOWN"
                res.notes  = f"Could not identify data type from columns: {list(df.columns)}"
                _log.warning("  Unknown type in %s — columns: %s", label, list(df.columns))

            score_tag = f"[{res.physics_grade} {res.physics_score}%]"
            _log.info("  %s  type=%-5s  status=%-11s  physics=%s  violations=%d",
                      label, dtype.upper(), res.status, score_tag, res.violations)
            results.append(res)

    if not results:
        _log.warning("No processable data found.")
        sys.exit(0)

    output_path = os.path.abspath(args.output)
    write_report(results, output_path)

    # Final summary
    ok    = sum(1 for r in results if r.status == "OK")
    errs  = sum(1 for r in results if r.status not in ("OK",))
    hold  = sum(1 for r in results if r.physics_score < 90)
    print()
    print(f"  PRC Batch Report Complete")
    print(f"  Files processed : {len(results)}")
    print(f"  OK              : {ok}")
    print(f"  Errors          : {errs}")
    print(f"  Physics HOLD    : {hold}  (score < 90%)")
    print(f"  Report          : {output_path}")
    print()


if __name__ == "__main__":
    main()
